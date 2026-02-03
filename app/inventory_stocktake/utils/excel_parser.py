from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List

from openpyxl import load_workbook

from .validators import (
    ValidationError,
    parse_non_negative_decimal,
    parse_non_negative_int,
    require_non_empty,
    validate_operation_type,
    validate_product_image,
    validate_status,
)

SHEET_NAME = "产品信息"

EXPECTED_HEADERS = [
    "操作类型",
    "物料编码",
    "中文名称",
    "泰文名称",
    "规格型号",
    "每组数量",
    "每件单价（泰铢）",
    "每组单价（泰铢）",
    "物料类别",
    "安全库存",
    "状态",
    "产品图片",
    "备注",
]


@dataclass
class ParsedMaterialRow:
    row_number: int
    data: Dict[str, Any]


def _normalize(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def parse_material_template(file_path: str) -> List[ParsedMaterialRow]:
    """解析 inventory_stocktake_template.xlsx。

    仅负责解析与字段级校验；“新增/更新”的业务规则在 service 层处理。

    Raises:
        ValidationError: 模板格式或字段校验失败。
    """

    wb = load_workbook(file_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValidationError(f"模板工作表名称必须为：{SHEET_NAME}")

    ws = wb[SHEET_NAME]

    header_row = [
        _normalize(c.value) for c in ws[1]
    ]
    if header_row[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise ValidationError("模板表头不匹配，请使用标准模板")

    rows: List[ParsedMaterialRow] = []

    for idx in range(2, ws.max_row + 1):
        row_values = {EXPECTED_HEADERS[i]: ws.cell(row=idx, column=i + 1).value for i in range(len(EXPECTED_HEADERS))}
        # 空行跳过
        if all(_normalize(v) == "" for v in row_values.values()):
            continue

        try:
            op_type = validate_operation_type(row_values["操作类型"])
            material_code = require_non_empty(row_values["物料编码"], "物料编码")
            cn_name = require_non_empty(row_values["中文名称"], "中文名称")
            th_name = _normalize(row_values["泰文名称"]) or None
            spec_model = require_non_empty(row_values["规格型号"], "规格型号")
            per_group_qty = parse_non_negative_int(row_values["每组数量"], "每组数量")
            price_per_case = parse_non_negative_decimal(row_values["每件单价（泰铢）"], "每件单价（泰铢）")
            price_per_group = parse_non_negative_decimal(row_values["每组单价（泰铢）"], "每组单价（泰铢）")
            category = require_non_empty(row_values["物料类别"], "物料类别")

            safety_raw = row_values["安全库存"]
            safety_stock = None
            if _normalize(safety_raw) != "":
                safety_stock = parse_non_negative_int(safety_raw, "安全库存")

            status = validate_status(row_values["状态"])
            product_image = validate_product_image(row_values["产品图片"])
            remark = _normalize(row_values["备注"]) or None

            rows.append(
                ParsedMaterialRow(
                    row_number=idx,
                    data={
                        "operation_type": op_type,
                        "material_code": material_code,
                        "cn_name": cn_name,
                        "th_name": th_name,
                        "spec_model": spec_model,
                        "per_group_qty": per_group_qty,
                        "price_per_case": price_per_case,
                        "price_per_group": price_per_group,
                        "category": category,
                        "safety_stock": safety_stock,
                        "status": status,
                        "product_image": product_image,
                        "remark": remark,
                    },
                )
            )
        except ValidationError:
            raise
        except Exception as e:
            raise ValidationError(f"第{idx}行解析失败：{e}") from e

    return rows
